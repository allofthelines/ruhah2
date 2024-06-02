'''
- ITEM.ITEMID == 123123
- EXCEL.ITEMSTAGS 123123    SHIRT

RUN SCRIPT

- ITEM.TAGS ++ SHIRT

'''

import openpyxl
from django.core.management.base import BaseCommand
from studio.models import Item  # Import the Item model

class Command(BaseCommand):
    help = 'Write tags to Item instances from an Excel file'
    
    def handle(self, *args, **kwargs):
        # Load the workbook and the specific sheet
        try:
            wb = openpyxl.load_workbook('tables.xlsx')
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR('The file tables.xlsx was not found.'))
            return
        
        sheet = wb['itemstags']
        
        # Get the headers
        headers = [sheet.cell(row=1, column=1).value, sheet.cell(row=1, column=2).value]
        if headers != ['item_id', 'tag_name']:
            self.stdout.write(self.style.ERROR('Headers do not match the expected values'))
            return
        
        # Read the data
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
            item_id = row[0].value  # First column (A)
            tag_name = row[1].value  # Second column (B)
            
            if item_id and tag_name:
                try:
                    item = Item.objects.get(itemid=item_id)
                    if item.tags:
                        tags = item.tags.split()
                        if tag_name not in tags:
                            item.tags += f" {tag_name}"
                            item.save()
                            self.stdout.write(self.style.SUCCESS(f'Successfully updated item {item_id} with tag: {tag_name}'))
                        else:
                            self.stdout.write(self.style.WARNING(f'Tag "{tag_name}" already exists for item {item_id}'))
                    else:
                        item.tags = tag_name
                        item.save()
                        self.stdout.write(self.style.SUCCESS(f'Successfully updated item {item_id} with tag: {tag_name}'))
                except Item.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f'Item with id {item_id} does not exist'))
                    