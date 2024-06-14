'''
- EXCEL.ITEMS.ITEMID == 123123
- ITEM.ITEMID != 123123

RUN SCRIPT

- ITEM.ITEMID = 123123
- ITEM.NAME = EXCEL.ITEMS.NAME
- ITEM.FIELD = EXCEL.ITEMS.FIELD

'''

import openpyxl
from django.core.management.base import BaseCommand
from studio.models import Item  # Import the Item model


class Command(BaseCommand):
    help = 'Write fields to Item instances from an Excel file'

    def handle(self, *args, **kwargs):
        # Load the workbook and the specific sheet
        try:
            wb = openpyxl.load_workbook('tables.xlsx')
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR('The file tables.xlsx was not found.'))
            return

        sheet = wb['items']


        # Get the headers
        headers = [sheet.cell(row=1, column=1).value,
                   sheet.cell(row=1, column=2).value,
                   sheet.cell(row=1, column=3).value,
                   sheet.cell(row=1, column=4).value,
                   sheet.cell(row=1, column=5).value,
                   sheet.cell(row=1, column=6).value,
                   sheet.cell(row=1, column=9).value,
                   sheet.cell(row=1, column=11).value]

        expected_headers = ['itemid', 'name', 'type', 'brand', 'condition', 'location', 'cat', 'price']
        if headers != expected_headers:
            print(headers)
            self.stdout.write(self.style.ERROR('Headers do not match the expected values'))
            return

        # Read the data
        for row in sheet.iter_rows(min_row=3, min_col=1, max_col=11):
            itemid = row[0].value  # Second column (B)
            name = row[1].value  # Third column (C)
            type = row[2].value # Fourth column (D)
            brand = row[3].value  # Fifth column (E)
            condition = row[4].value  # Sixth column (F)
            location = row[5].value  # Seventh column (G)
            cat = row[8].value  # Tenth column (J)
            price = row[10].value

            if itemid:
                try:
                    item = Item.objects.get(itemid=itemid)
                    self.stdout.write(self.style.WARNING(f'Item with id {itemid} already exists'))
                except Item.DoesNotExist:
                    item = Item(
                        itemid=itemid,
                        name=name,
                        brand=condition,
                        condition=condition,
                        location=location,
                        cat=cat,
                        size_xyz=size_xyz,
                        price=price
                    )
                    # Append the category to the tags field
                    if item.tags:
                        item.tags += f" {cat}"
                    else:
                        item.tags = cat
                    item.save()
                    self.stdout.write(self.style.SUCCESS(f'Successfully created item with id: {itemid}'))
